from __future__ import annotations

import json
import re
import shutil
import time
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import urljoin
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import requests
import xlrd
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from playwright.sync_api import (
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)


# -----------------------------------------------------------------------------
# تنظیمات
# -----------------------------------------------------------------------------
MDP_URL = "https://mdp.ihio.gov.ir/"
ESATA_PAGE_URL = "https://esata.ir/web/sakhad/drug/"
TAMIN_URL = (
    "https://darman.tamin.ir/Forms/Public/"
    "Druglist.aspx?pagename=hdpDrugList"
)

OUTPUT_DIR = Path("output")
TIMEOUT_SECONDS = 180
TIMEOUT_MS = TIMEOUT_SECONDS * 1000
TAMIN_PAGE_DELAY_SECONDS = 0.25
MAX_RETRIES = 3
VERIFY_TLS = True

# اگر True باشد، خرابی یک منبع مانع ساخت فایل نهایی نمی‌شود و در شیت آن
# منبع، پیام خطا قرار می‌گیرد.
# GLYMIZE publishes only complete source runs. A failed source must keep the
# previous healthy catalogue active and create an admin notification.
CONTINUE_ON_ERROR = False

GRID_NAME = "ctl00_ContentPlaceHolder1_Grd_Dr"
GRID_PREFIX = GRID_NAME
SEARCH_BUTTON = "#ctl00_ContentPlaceHolder1_btnSearch"

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/150.0.0.0 Safari/537.36"
)


def tehran_now() -> datetime:
    """Return Tehran time even on a fresh Windows Python without system tzdata."""
    try:
        return datetime.now(ZoneInfo("Asia/Tehran"))
    except ZoneInfoNotFoundError:
        # Iran has used UTC+03:30 year-round since 22 September 2022. The
        # fallback keeps a current extraction operational; installing tzdata
        # remains the preferred path for complete historical timezone data.
        return datetime.now(timezone(timedelta(hours=3, minutes=30)))

# -----------------------------------------------------------------------------
# ابزارهای عمومی
# -----------------------------------------------------------------------------
def new_http_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept-Language": "fa-IR,fa;q=0.9,en;q=0.7",
        }
    )
    return session


def english_digits(value: str) -> str:
    table = str.maketrans(
        "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
        "01234567890123456789",
    )
    return value.translate(table)


def clean_cell(value: Any) -> Any:
    if value is None:
        return ""

    if isinstance(value, (int, float, bool)):
        return value

    text = str(value)
    # حذف کاراکترهای کنترلی نامعتبر در XML اکسل، به‌جز tab/newline.
    return re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", text)


def normalize_matrix(rows: list[list[Any]]) -> list[list[Any]]:
    cleaned = [[clean_cell(cell) for cell in row] for row in rows]

    # حذف ردیف‌های کاملاً خالی ابتدا و انتهای فایل.
    while cleaned and not any(str(cell).strip() for cell in cleaned[0]):
        cleaned.pop(0)
    while cleaned and not any(str(cell).strip() for cell in cleaned[-1]):
        cleaned.pop()

    if not cleaned:
        return [["پیام"], ["هیچ داده‌ای دریافت نشد."]]

    last_nonempty_column = 0
    for row in cleaned:
        for index, cell in enumerate(row):
            if str(cell).strip():
                last_nonempty_column = max(last_nonempty_column, index + 1)

    last_nonempty_column = max(last_nonempty_column, 1)

    normalized: list[list[Any]] = []
    for row in cleaned:
        current = row[:last_nonempty_column]
        if len(current) < last_nonempty_column:
            current += [""] * (last_nonempty_column - len(current))
        normalized.append(current)

    return normalized


def deduplicate_rows(rows: list[list[Any]]) -> list[list[Any]]:
    if len(rows) <= 2:
        return rows

    result = [rows[0]]
    seen: set[tuple[str, ...]] = set()

    for row in rows[1:]:
        key = tuple(str(value).strip() for value in row)
        if key in seen:
            continue
        seen.add(key)
        result.append(row)

    return result


def error_sheet(source_url: str, exc: Exception) -> list[list[Any]]:
    return [
        ["وضعیت", "خطا در دریافت اطلاعات"],
        ["منبع", source_url],
        ["نوع خطا", type(exc).__name__],
        ["شرح", str(exc)],
    ]


# -----------------------------------------------------------------------------
# منبع اول: بیمه سلامت — دانلود Excel با POST
# -----------------------------------------------------------------------------
def build_mdp_payload(initial_html: str) -> tuple[dict[str, str], dict[str, tuple]]:
    soup = BeautifulSoup(initial_html, "html.parser")

    # هر hidden input موجود در نسخه فعلی صفحه نیز همراه درخواست ارسال می‌شود.
    data: dict[str, str] = {}
    for element in soup.select("input[type='hidden'][name]"):
        data[element.get("name", "")] = element.get("value", "")

    service_state = json.dumps(
        [{"value": "842", "text": "داروخانه", "index": 0}],
        ensure_ascii=True,
        separators=(",", ":"),
    )
    change_state = json.dumps(
        [{"value": "768", "text": "آخرين تغييرات", "index": 2}],
        ensure_ascii=True,
        separators=(",", ":"),
    )
    page_size_state = json.dumps(
        [{"value": "10", "text": "10", "index": 0}],
        ensure_ascii=True,
        separators=(",", ":"),
    )

    filter_values = {
        "cmbServiceType": "842",
        "_cmbServiceType_state": service_state,
        "txtSrchSrvCode": "",
        "txtSrchSrvName": "",
        "txtSrchChgDate": "",
        "txtSrchChgDate2": "",
        "txtSerSize": "",
        "txtSrchFeranshiz": "",
        "cmbSrchSrvChgStatus": "768",
        "_cmbSrchSrvChgStatus_state": change_state,
        "cmbDeviceGroup": "",
        "_cmbDeviceGroup_state": "",
        "txtSerchUniversal": "",
        "cmbSrchFehrest": "",
        "_cmbSrchFehrest_state": "",
        "cmbSrchExpert": "",
        "_cmbSrchExpert_state": "",
        "cmbSrchRes": "",
        "_cmbSrchRes_state": "",
        "cmbSrchSpc": "",
        "_cmbSrchSpc_state": "",
        "cmbBimeProdType": "",
        "_cmbBimeProdType_state": "",
    }

    direct_event_config = {
        "config": {
            "extraParams": {
                "values": json.dumps(
                    filter_values,
                    ensure_ascii=False,
                    separators=(",", ":"),
                )
            }
        }
    }

    empty_fields = [
        "txtSrchSrvCode",
        "txtSrchSrvName",
        "txtSrchChgDate",
        "txtSrchChgDate2",
        "txtSerSize",
        "txtSrchFeranshiz",
        "cmbDeviceGroup",
        "_cmbDeviceGroup_state",
        "txtSerchUniversal",
        "cmbSrchFehrest",
        "_cmbSrchFehrest_state",
        "cmbSrchExpert",
        "_cmbSrchExpert_state",
        "cmbSrchRes",
        "_cmbSrchRes_state",
        "cmbSrchSpc",
        "_cmbSrchSpc_state",
        "cmbBimeProdType",
        "_cmbBimeProdType_state",
        "grdSearchServiceRowMod",
        "ESALAT_REQ",
        "_ESALAT_REQ_state",
        "SRV_GUID",
        "SRV_CODE_USER",
        "SRV_CODE_BRAND",
        "SRV_CODE_FDA",
        "SRV_RXCODE",
        "SRV_NAME_EN",
        "SRV_DESC_OLD",
        "SRV_DOZE",
        "SRV_SHP_FK",
        "_SRV_SHP_FK_state",
        "SRV_FILE_DESC",
        "SRV_STATUS",
        "_SRV_STATUS_state",
        "SRV_STATUS_CHG_DATE",
        "ucHelp_cmbHelpType",
        "_ucHelp_cmbHelpType_state",
        "ucServiceDoc_cmbDocType",
        "_ucServiceDoc_cmbDocType_state",
        "ucFdaTdCompare_txtSrchSrvCode",
        "ucFdaTdCompare_cmbOnlyDifferencePrice",
        "_ucFdaTdCompare_cmbOnlyDifferencePrice_state",
    ]

    for field in empty_fields:
        data.setdefault(field, "")

    data.update(
        {
            "cmbServiceType": "داروخانه",
            "_cmbServiceType_state": service_state,
            "cmbSrchSrvChgStatus": "آخرين تغييرات",
            "_cmbSrchSrvChgStatus_state": change_state,
            "ComboBox1": "10",
            "_ComboBox1_state": page_size_state,
            "__VIEWSTATEGENERATOR": data.get(
                "__VIEWSTATEGENERATOR",
                "3CD35005",
            ),
            "__EVENTTARGET": "ResourceManager1",
            "__EVENTARGUMENT": "btnExcelExport|event|Click",
            "submitDirectEventConfig": json.dumps(
                direct_event_config,
                ensure_ascii=False,
                separators=(",", ":"),
            ),
            "__ExtNetDirectEventMarker": "delta=true",
        }
    )

    # وجود فایل خالی باعث می‌شود requests درخواست را multipart/form-data ارسال کند.
    files = {
        "ucServiceDoc_fileDocument": (
            "",
            b"",
            "application/octet-stream",
        )
    }

    return data, files


def download_mdp_excel() -> bytes:
    print("[1/3] بیمه سلامت: در حال دریافت آخرین فایل Excel...")

    with new_http_session() as session:
        initial = session.get(
            MDP_URL,
            timeout=TIMEOUT_SECONDS,
            verify=VERIFY_TLS,
        )
        initial.raise_for_status()

        data, files = build_mdp_payload(initial.text)

        response = session.post(
            MDP_URL,
            data=data,
            files=files,
            headers={
                "Origin": "https://mdp.ihio.gov.ir",
                "Referer": MDP_URL,
            },
            timeout=TIMEOUT_SECONDS,
            verify=VERIFY_TLS,
        )
        response.raise_for_status()

    content_type = response.headers.get("Content-Type", "").lower()
    raw = response.content

    # فایل xls قدیمی باید با امضای OLE/CFBF شروع شود.
    is_ole_xls = raw.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1")
    if not is_ole_xls and "excel" not in content_type and "msexcel" not in content_type:
        preview = raw[:500].decode("utf-8", errors="replace")
        raise RuntimeError(
            "پاسخ بیمه سلامت فایل Excel نبود. "
            f"Content-Type={content_type!r}; preview={preview!r}"
        )

    print(f"      دریافت شد: {len(raw):,} بایت")
    return raw


# -----------------------------------------------------------------------------
# منبع دوم: ساتا — استخراج لینک آخرین فایل از صفحه
# -----------------------------------------------------------------------------
def extract_esata_file_url(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")

    config_element = soup.find("script", id="excelViewerConfig")
    if config_element is not None:
        try:
            config = json.loads(config_element.get_text(strip=True))
            file_url = config.get("fileUrl")
            if file_url:
                return urljoin(ESATA_PAGE_URL, file_url)
        except json.JSONDecodeError:
            pass

    patterns = [
        r'"fileUrl"\s*:\s*"([^"]+\.xlsx(?:\?[^\"]*)?)"',
        r"['\"]([^'\"]+\.xlsx(?:\?[^'\"]*)?)['\"]",
    ]
    for pattern in patterns:
        match = re.search(pattern, html, flags=re.IGNORECASE)
        if match:
            return urljoin(ESATA_PAGE_URL, match.group(1).replace("\\/", "/"))

    raise RuntimeError("لینک فایل xlsx در صفحه ساتا پیدا نشد.")


def download_esata_excel() -> tuple[bytes, str]:
    print("[2/3] ساتا: در حال پیدا کردن و دریافت آخرین فایل Excel...")

    with new_http_session() as session:
        page_response = session.get(
            ESATA_PAGE_URL,
            timeout=TIMEOUT_SECONDS,
            verify=VERIFY_TLS,
        )
        page_response.raise_for_status()

        download_url = extract_esata_file_url(page_response.text)
        print(f"      لینک فعلی: {download_url}")

        file_response = session.get(
            download_url,
            headers={"Referer": ESATA_PAGE_URL},
            timeout=TIMEOUT_SECONDS,
            verify=VERIFY_TLS,
        )
        file_response.raise_for_status()

    raw = file_response.content
    content_type = file_response.headers.get("Content-Type", "").lower()

    if not raw.startswith(b"PK\x03\x04") and "spreadsheet" not in content_type:
        preview = raw[:500].decode("utf-8", errors="replace")
        raise RuntimeError(
            "پاسخ ساتا فایل xlsx نبود. "
            f"Content-Type={content_type!r}; preview={preview!r}"
        )

    print(f"      دریافت شد: {len(raw):,} بایت")
    return raw, download_url


# -----------------------------------------------------------------------------
# خواندن xls/xlsx با وابستگی‌های پین‌شده محلی
# -----------------------------------------------------------------------------
def parse_excel_bytes(raw: bytes) -> list[list[Any]]:
    sheets: list[list[list[Any]]] = []
    if raw.startswith(b"PK\x03\x04"):
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            sheets.append([["" if value is None else value for value in row] for row in worksheet.iter_rows(values_only=True)])
    else:
        workbook = xlrd.open_workbook(file_contents=raw)
        for worksheet in workbook.sheets():
            sheets.append([[worksheet.cell_value(row, column) for column in range(worksheet.ncols)] for row in range(worksheet.nrows)])
    if not sheets:
        raise RuntimeError("فایل Excel هیچ شیتی نداشت.")
    best = max(sheets, key=lambda rows: sum(len(row) for row in rows))
    return normalize_matrix(best)


# -----------------------------------------------------------------------------
# منبع سوم: تأمین اجتماعی — پیمایش همه صفحات DevExpress
# -----------------------------------------------------------------------------
def wait_for_tamin_grid(page: Page) -> None:
    page.wait_for_function(
        """
        gridName => {
          const grid = window[gridName];
          return grid &&
                 typeof grid.GetPageCount === "function" &&
                 document.querySelector(`#${gridName}_DXMainTable`) &&
                 !grid.InCallback();
        }
        """,
        arg=GRID_NAME,
        timeout=TIMEOUT_MS,
    )


def tamin_reported_total(page: Page) -> int | None:
    selector = f"#{GRID_PREFIX}_DXPagerBottom .dxpSummary_Office2003Blue"
    if page.locator(selector).count() == 0:
        return None

    text = english_digits(page.locator(selector).inner_text())
    match = re.search(r"\(([\d,]+)\s*ردیف\)", text)
    if match is None:
        return None

    return int(match.group(1).replace(",", ""))


def tamin_goto_page(page: Page, page_index: int) -> None:
    current_page = int(
        page.evaluate(
            "gridName => window[gridName].GetPageIndex()",
            GRID_NAME,
        )
    )
    if current_page == page_index:
        return

    page.evaluate(
        """
        ({gridName, pageIndex}) => {
          window[gridName].GotoPage(pageIndex);
        }
        """,
        {"gridName": GRID_NAME, "pageIndex": page_index},
    )

    page.wait_for_function(
        """
        ({gridName, pageIndex}) => {
          const grid = window[gridName];
          return grid &&
                 grid.GetPageIndex() === pageIndex &&
                 !grid.InCallback();
        }
        """,
        arg={"gridName": GRID_NAME, "pageIndex": page_index},
        timeout=TIMEOUT_MS,
    )


def tamin_headers(page: Page) -> list[str]:
    headers = page.eval_on_selector_all(
        f"#{GRID_PREFIX}_DXHeadersRow0 > td",
        """
        cells => cells
          .map(cell => cell.innerText.trim())
          .filter(text => text.length > 0)
        """,
    )
    if not headers:
        raise RuntimeError("ستون‌های جدول تأمین اجتماعی پیدا نشدند.")
    return headers


def tamin_rows(page: Page, column_count: int) -> list[list[Any]]:
    rows = page.eval_on_selector_all(
        f"tr[id^='{GRID_PREFIX}_DXDataRow']",
        """
        rows => rows.map(row =>
          Array.from(row.children).map(cell => cell.innerText.trim())
        )
        """,
    )

    result: list[list[Any]] = []
    for row in rows:
        current = row[:column_count]
        if len(current) < column_count:
            current += [""] * (column_count - len(current))
        result.append(current)
    return result


def scrape_tamin(context: BrowserContext) -> tuple[list[list[Any]], int | None]:
    print("[3/3] تأمین اجتماعی: در حال استخراج همه صفحات جدول...")

    page = context.new_page()
    page.set_default_timeout(TIMEOUT_MS)

    try:
        page.goto(
            TAMIN_URL,
            wait_until="domcontentloaded",
            timeout=TIMEOUT_MS,
        )
        wait_for_tamin_grid(page)

        page.locator(SEARCH_BUTTON).click()
        wait_for_tamin_grid(page)

        page_count = int(
            page.evaluate(
                "gridName => window[gridName].GetPageCount()",
                GRID_NAME,
            )
        )
        reported_total = tamin_reported_total(page)
        headers = tamin_headers(page)
        all_rows: list[list[Any]] = []

        print(f"      تعداد صفحات فعلی: {page_count:,}")
        if reported_total is not None:
            print(f"      تعداد ردیف اعلام‌شده: {reported_total:,}")

        for page_index in range(page_count):
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    tamin_goto_page(page, page_index)
                    current_rows = tamin_rows(page, len(headers))
                    if not current_rows:
                        raise RuntimeError("صفحه جدول بدون ردیف برگشت.")

                    all_rows.extend(current_rows)
                    print(
                        f"      صفحه {page_index + 1:,}/{page_count:,}"
                        f" — ردیف‌های جمع‌آوری‌شده: {len(all_rows):,}",
                        end="\r",
                        flush=True,
                    )
                    break

                except (PlaywrightTimeoutError, RuntimeError) as exc:
                    if attempt == MAX_RETRIES:
                        raise RuntimeError(
                            f"دریافت صفحه {page_index + 1} پس از "
                            f"{MAX_RETRIES} تلاش ناموفق بود."
                        ) from exc
                    time.sleep(2)

            time.sleep(TAMIN_PAGE_DELAY_SECONDS)

        print()
        matrix = normalize_matrix([headers, *all_rows])
        matrix = deduplicate_rows(matrix)
        print(f"      ردیف نهایی: {max(len(matrix) - 1, 0):,}")
        return matrix, reported_total

    finally:
        page.close()


# -----------------------------------------------------------------------------
# ساخت فایل نهایی سه‌شیتی با openpyxl
# -----------------------------------------------------------------------------
def estimate_widths(rows: list[list[Any]]) -> list[int]:
    column_count = max((len(row) for row in rows), default=1)
    widths: list[int] = []

    # برای سرعت، حداکثر 500 ردیف اول بررسی می‌شود.
    sample = rows[:500]
    for column_index in range(column_count):
        lengths = []
        for row in sample:
            value = row[column_index] if column_index < len(row) else ""
            lengths.append(len(str(value)))
        widths.append(min(max(max(lengths, default=8) + 2, 10), 50))

    return widths


def write_combined_workbook(
    sheets: list[dict[str, Any]],
    output_path: Path,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.title = "خروجی تجمیعی اطلاعات دارویی"
    workbook.properties.subject = "سه منبع بیمه سلامت، ساتا و تأمین اجتماعی"
    workbook.properties.creator = "GLYMIZE Python Runner"
    for sheet in sheets:
        rows = normalize_matrix(sheet["rows"])
        worksheet = workbook.create_sheet(sheet["name"][:31])
        for row in rows:
            worksheet.append(row)
        widths = estimate_widths(rows)
        for column_index, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(column_index)].width = width
        if rows and rows[0]:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(rows[0]))}{len(rows)}"
        worksheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def fetch_source(
    name: str,
    source_url: str,
    function,
) -> tuple[list[list[Any]], dict[str, Any]]:
    started = time.monotonic()
    try:
        result = function()
        elapsed = time.monotonic() - started
        return result, {
            "name": name,
            "source": source_url,
            "status": "ok",
            "seconds": round(elapsed, 2),
        }
    except Exception as exc:
        if not CONTINUE_ON_ERROR:
            raise
        print(f"      خطا در {name}: {exc}")
        elapsed = time.monotonic() - started
        return error_sheet(source_url, exc), {
            "name": name,
            "source": source_url,
            "status": "error",
            "error": str(exc),
            "seconds": round(elapsed, 2),
        }


# -----------------------------------------------------------------------------
# اجرای اصلی
# -----------------------------------------------------------------------------
def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    now = tehran_now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    archive_path = OUTPUT_DIR / f"drug_sources_{timestamp}.xlsx"
    latest_path = OUTPUT_DIR / "drug_sources_latest.xlsx"

    metadata: list[dict[str, Any]] = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(
            ignore_https_errors=not VERIFY_TLS,
            locale="fa-IR",
            user_agent=USER_AGENT,
        )

        mdp_rows, mdp_meta = fetch_source(
            "بیمه سلامت",
            MDP_URL,
            lambda: parse_excel_bytes(download_mdp_excel()),
        )
        metadata.append(mdp_meta)

        def esata_job() -> list[list[Any]]:
            raw, _download_url = download_esata_excel()
            return parse_excel_bytes(raw)

        esata_rows, esata_meta = fetch_source(
            "ساتا",
            ESATA_PAGE_URL,
            esata_job,
        )
        metadata.append(esata_meta)

        tamin_rows_result, tamin_meta = fetch_source(
            "تأمین اجتماعی",
            TAMIN_URL,
            lambda: scrape_tamin(context)[0],
        )
        metadata.append(tamin_meta)

        sheets = [
            {"name": "بیمه سلامت", "rows": mdp_rows},
            {"name": "ساتا", "rows": esata_rows},
            {"name": "تامین اجتماعی", "rows": tamin_rows_result},
        ]

        print("در حال ساخت فایل نهایی سه‌شیتی...")
        write_combined_workbook(sheets, archive_path)

        context.close()
        browser.close()

    shutil.copyfile(archive_path, latest_path)

    report_path = OUTPUT_DIR / "drug_sources_latest_report.json"
    report_path.write_text(
        json.dumps(
            {
                "created_at": now.isoformat(timespec="seconds"),
                "archive_file": str(archive_path),
                "latest_file": str(latest_path),
                "sources": metadata,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    print()
    print(f"فایل آرشیوی: {archive_path.resolve()}")
    print(f"آخرین نسخه:   {latest_path.resolve()}")
    print(f"گزارش اجرا:   {report_path.resolve()}")


if __name__ == "__main__":
    main()
