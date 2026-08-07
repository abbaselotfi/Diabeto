from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from normalize_bundle_runtime import build_bundle, coverage_payload


class RuntimeCoverageFallbackTests(unittest.TestCase):
    def test_health_coverage_is_derived_from_official_amounts_when_percent_is_blank(self) -> None:
        headers = [
            "کد ژنريک",
            "کد برند",
            "عنوان",
            " درصد سهم سازمان با احتساب يارانه ارزي",
            " مبلغ سهم سازمان با احتساب يارانه ارزي",
            "قيمت کل مورد درتعهد با احتساب يارانه ارزي",
            "سهم سازمان",
        ]
        row = {
            "کد ژنريک": "00001",
            "کد برند": "00001",
            "عنوان": "A.C.A",
            " درصد سهم سازمان با احتساب يارانه ارزي": None,
            " مبلغ سهم سازمان با احتساب يارانه ارزي": "22,225",
            "قيمت کل مورد درتعهد با احتساب يارانه ارزي": "31,000",
            "سهم سازمان": "70%",
        }
        coverage, warning = coverage_payload(
            row,
            headers,
            "health_insurance",
            "https://mdp.ihio.gov.ir/",
            "2026-08-07T00:00:00+00:00",
        )
        self.assertIsNone(warning)
        self.assertIsNotNone(coverage)
        assert coverage is not None
        self.assertAlmostEqual(coverage["percent"], 71.693548, places=6)
        self.assertEqual(coverage["genericCode"], "00001")
        self.assertIn("محاسبه", coverage["sourceReference"])

    def test_health_explicit_share_percent_is_safe_last_fallback(self) -> None:
        headers = [
            "کد ژنريک",
            "عنوان",
            " درصد سهم سازمان با احتساب يارانه ارزي",
            " مبلغ سهم سازمان با احتساب يارانه ارزي",
            "قيمت کل مورد درتعهد با احتساب يارانه ارزي",
            "سهم سازمان",
        ]
        row = {
            "کد ژنريک": "02015",
            "عنوان": "ACARBOSE",
            " درصد سهم سازمان با احتساب يارانه ارزي": " ",
            " مبلغ سهم سازمان با احتساب يارانه ارزي": None,
            "قيمت کل مورد درتعهد با احتساب يارانه ارزي": None,
            "سهم سازمان": "70%",
        }
        coverage, warning = coverage_payload(
            row,
            headers,
            "health_insurance",
            "https://mdp.ihio.gov.ir/",
            "2026-08-07T00:00:00+00:00",
        )
        self.assertIsNone(warning)
        self.assertIsNotNone(coverage)
        assert coverage is not None
        self.assertEqual(coverage["percent"], 70)
        self.assertEqual(coverage["sourceReference"], "سهم سازمان")

    def test_plain_numeric_share_is_not_misread_as_percent(self) -> None:
        headers = [
            "کد ژنريک",
            "عنوان",
            " درصد سهم سازمان با احتساب يارانه ارزي",
            " مبلغ سهم سازمان با احتساب يارانه ارزي",
            "قيمت کل مورد درتعهد با احتساب يارانه ارزي",
            "سهم سازمان",
        ]
        row = {
            "کد ژنريک": "02015",
            "عنوان": "ACARBOSE",
            " درصد سهم سازمان با احتساب يارانه ارزي": None,
            " مبلغ سهم سازمان با احتساب يارانه ارزي": None,
            "قيمت کل مورد درتعهد با احتساب يارانه ارزي": None,
            "سهم سازمان": "70000",
        }
        coverage, warning = coverage_payload(
            row,
            headers,
            "health_insurance",
            "https://mdp.ihio.gov.ir/",
            "2026-08-07T00:00:00+00:00",
        )
        self.assertIsNone(coverage)
        self.assertIsNotNone(warning)
        assert warning is not None
        self.assertIn("nonblocking-ihio-missing-percent", warning)

    def test_missing_health_percent_does_not_mark_entire_source_incomplete(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            scope_path = root / "scope.json"
            scope_path.write_text(
                json.dumps({
                    "schemaVersion": 1,
                    "entries": [
                        {
                            "canonicalName": "Acarbose",
                            "aliases": ["Acarbose", "ACARBOSE"],
                            "clinicalDomains": ["diabetes"],
                        }
                    ],
                }),
                encoding="utf-8",
            )

            nfi_path = root / "nfi.xlsx"
            nfi = Workbook()
            sheet = nfi.active
            sheet.title = "NFI_Products_1"
            sheet.append([
                "Generic_INN",
                "NFI_Detail_ID",
                "NFI_Title_FA",
                "NFI_Generic_Code",
                "NFI_IRC_Code",
                "Dosage_Form",
                "Strength",
                "Match_Confidence_0_100",
            ])
            sheet.append([
                "Acarbose",
                "1",
                "ACARBOSE TEST TABLET 100 mg",
                "02015",
                "6260001",
                "TABLET",
                "100 mg",
                99,
            ])
            nfi.save(nfi_path)

            insurance_path = root / "insurance.xlsx"
            insurance = Workbook()
            health = insurance.active
            health.title = "بیمه سلامت"
            health.append([
                "کد ژنريک",
                "کد برند",
                "عنوان",
                " درصد سهم سازمان با احتساب يارانه ارزي",
                " مبلغ سهم سازمان با احتساب يارانه ارزي",
                "قيمت کل مورد درتعهد با احتساب يارانه ارزي",
                "سهم سازمان",
            ])
            health.append(["02015", "02015", "ACARBOSE", None, None, None, "70000"])

            armed = insurance.create_sheet("ساتا")
            armed.append([
                "کد ژنريک",
                "نام ژنريك دارو",
                "درصد سهم بيمار عادي با احتساب يارانه",
            ])
            armed.append(["02015", "ACARBOSE 100 MG TABLET ORAL", 0.13])

            social = insurance.create_sheet("تامین اجتماعی")
            social.append([
                "کد دارو",
                "نام دارو",
                "درصد سازمان از قیمت بدون یارانه",
            ])
            social.append(["02015", "ACARBOSE 100 MG TABLET ORAL", "70%"])
            insurance.save(insurance_path)

            bundle = build_bundle(nfi_path, insurance_path, None, scope_path)
            sources = {item["sourceId"]: item for item in bundle["run"]["sources"]}

            self.assertEqual(sources["health_insurance"]["status"], "succeeded")
            self.assertEqual(sources["health_insurance"]["warningCount"], 1)
            self.assertEqual(sources["health_insurance"]["skippedCoverageRowCount"], 1)
            self.assertIsNone(sources["health_insurance"]["error"])
            self.assertIn("کنار گذاشته شد", sources["health_insurance"]["warning"])
            self.assertEqual(bundle["run"]["summary"]["errorCount"], 0)
            self.assertEqual(bundle["run"]["summary"]["warningCount"], 1)
            self.assertEqual(bundle["run"]["status"], "ready_to_publish")

    def test_authoritative_nfi_generic_overrides_seed_query_and_upgrades_insurance_identity(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            scope_path = root / "scope.json"
            scope_path.write_text(
                json.dumps({
                    "schemaVersion": 1,
                    "entries": [
                        {"canonicalName": "Gliclazide", "aliases": ["Gliclazide"], "clinicalDomains": ["diabetes"]},
                        {"canonicalName": "Metformin + gliclazide", "aliases": ["Metformin + gliclazide"], "clinicalDomains": ["diabetes"]},
                    ],
                }),
                encoding="utf-8",
            )

            nfi_path = root / "nfi.xlsx"
            nfi = Workbook()
            sheet = nfi.active
            sheet.title = "NFI_Products_1"
            sheet.append([
                "Generic_INN", "NFI_Generic_Name", "NFI_Detail_ID", "NFI_Title_FA",
                "NFI_Generic_Code", "NFI_IRC_Code", "Dosage_Form", "Strength",
                "Match_Confidence_0_100",
            ])
            sheet.append([
                "Metformin + gliclazide", "GLICLAZIDE", "101", "گلی کلازید قرص 80 mg",
                "1523", "6857265833682812", "TABLET", "80 mg", 73,
            ])
            nfi.save(nfi_path)

            insurance_path = root / "insurance.xlsx"
            insurance = Workbook()
            health = insurance.active
            health.title = "بیمه سلامت"
            health.append(["کد ژنريک", "کد برند", "عنوان", " درصد سهم سازمان با احتساب يارانه ارزي"])
            health.append(["01523", "01523", "GLICLAZIDE", "70%"])
            armed = insurance.create_sheet("ساتا")
            armed.append(["کد ژنريک", "نام ژنريك دارو", "درصد سهم بيمار عادي با احتساب يارانه"])
            armed.append(["01523", "GLICLAZIDE 80 MG TABLET ORAL", 0.3])
            social = insurance.create_sheet("تامین اجتماعی")
            social.append(["کد دارو", "نام دارو", "درصد سازمان از قیمت بدون یارانه"])
            social.append(["01523", "GLICLAZIDE 80 MG TABLET ORAL", "70%"])
            insurance.save(insurance_path)

            bundle = build_bundle(nfi_path, insurance_path, None, scope_path)
            nfi_record = next(record for record in bundle["records"] if record.get("brandRegistryCode") == "6857265833682812")
            self.assertEqual(nfi_record["genericName"], "Gliclazide")
            self.assertGreaterEqual(nfi_record["matchConfidence"], 0.99)
            insurer_records = [record for record in bundle["records"] if record["insuranceCoverages"]]
            self.assertEqual(len(insurer_records), 3)
            self.assertTrue(all(record["genericName"] == "Gliclazide" for record in insurer_records))
            self.assertTrue(all(record["matchConfidence"] >= 0.9 for record in insurer_records))
            self.assertEqual(bundle["run"]["summary"]["ambiguousMatchCount"], 0)
            self.assertEqual(bundle["run"]["status"], "ready_to_publish")

    def test_cross_insurer_code_consensus_corrects_nfi_when_authoritative_generic_column_is_missing(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            scope_path = root / "scope.json"
            scope_path.write_text(
                json.dumps({
                    "schemaVersion": 1,
                    "entries": [
                        {"canonicalName": "Gliclazide", "aliases": ["Gliclazide"], "clinicalDomains": ["diabetes"]},
                        {"canonicalName": "Metformin + gliclazide", "aliases": ["Metformin + gliclazide"], "clinicalDomains": ["diabetes"]},
                    ],
                }),
                encoding="utf-8",
            )

            nfi_path = root / "nfi.xlsx"
            nfi = Workbook()
            sheet = nfi.active
            sheet.title = "NFI_Products_1"
            sheet.append([
                "Generic_INN", "NFI_Detail_ID", "NFI_Title_FA", "NFI_Generic_Code",
                "NFI_IRC_Code", "Dosage_Form", "Strength", "Match_Confidence_0_100",
            ])
            sheet.append([
                "Metformin + gliclazide", "101", "دیابزید قرص 80 mg", "1523",
                "6857265833682812", "TABLET", "80 mg", 73,
            ])
            nfi.save(nfi_path)

            insurance_path = root / "insurance.xlsx"
            insurance = Workbook()
            health = insurance.active
            health.title = "بیمه سلامت"
            health.append(["کد ژنريک", "کد برند", "عنوان", " درصد سهم سازمان با احتساب يارانه ارزي"])
            health.append(["01523", "01523", "GLICLAZIDE", "70%"])
            armed = insurance.create_sheet("ساتا")
            armed.append(["کد ژنريک", "نام ژنريك دارو", "درصد سهم بيمار عادي با احتساب يارانه"])
            armed.append(["1523", "GLICLAZIDE 80 MG TABLET ORAL", 0.3])
            social = insurance.create_sheet("تامین اجتماعی")
            social.append(["کد دارو", "نام دارو", "درصد سازمان از قیمت بدون یارانه"])
            social.append(["001523", "GLICLAZIDE 80 MG TABLET ORAL", "70%"])
            insurance.save(insurance_path)

            bundle = build_bundle(nfi_path, insurance_path, None, scope_path)
            nfi_record = next(record for record in bundle["records"] if record.get("brandRegistryCode") == "6857265833682812")
            self.assertEqual(nfi_record["genericName"], "Gliclazide")
            self.assertGreaterEqual(nfi_record["matchConfidence"], 0.98)
            self.assertIn("insurer generic-code consensus 1523", nfi_record["sourceReference"])
            self.assertEqual(bundle["run"]["summary"]["ambiguousMatchCount"], 0)
            self.assertEqual(bundle["run"]["status"], "ready_to_publish")


if __name__ == "__main__":
    unittest.main()
