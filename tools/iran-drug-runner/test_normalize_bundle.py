from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from normalize_bundle import build_bundle, coverage_payload, number, price_payload


class NormalizeBundleTests(unittest.TestCase):
    @staticmethod
    def write_scope(path: Path) -> None:
        path.write_text(
            json.dumps({
                "schemaVersion": 1,
                "entries": [
                    {"canonicalName": "Acarbose", "aliases": ["Acarbose", "ACARBOSE"], "clinicalDomains": ["diabetes"]},
                    {"canonicalName": "Metformin", "aliases": ["Metformin", "METFORMIN"], "clinicalDomains": ["diabetes"]},
                ],
            }),
            encoding="utf-8",
        )

    @staticmethod
    def write_realistic_insurance_workbook(path: Path, generic_name: str = "Acarbose", generic_code: str = "02015") -> None:
        insurance = Workbook()
        generic_upper = generic_name.upper()

        health = insurance.active
        health.title = "بیمه سلامت"
        health.append([
            "رديف", "کد ژنريک", "کد برند", "عنوان", "شرايط تعهد", "شکل", "دوز",
            "يارانه ارزي", " مبلغ سهم سازمان با احتساب يارانه ارزي",
            " درصد سهم سازمان با احتساب يارانه ارزي", "قيمت",
        ])
        health.append([1, generic_code, generic_code, generic_upper, "توليد داخل ديابت", "TAB", "100 mg", "48700", "69,910", "88.493", "30,300"])

        armed = insurance.create_sheet("ساتا")
        armed.append([
            "کد ژنريک", "نام ژنريك دارو", "قيمت با احتساب يارانه",
            "يارانه ارزی دولت (دارویار)", "قيمت بدون احتساب يارانه",
            "درصد سهم بيمار عادي با احتساب يارانه",
        ])
        armed.append([generic_code, f"{generic_upper} 100 MG TABLET ORAL  TABLET  {generic_code}", 27100, 3793, 23307, 0.1290055350553505])

        social = insurance.create_sheet("تامین اجتماعی")
        social.append([
            "کد دارو", "نام دارو", "بیمه", "قیمت دارو بدون یارانه",
            "درصد سازمان از قیمت بدون یارانه", "یارانه دولت", "جمع مورد قبول سازمان",
        ])
        social.append([generic_code, f"{generic_upper} 100 MG TABLET ORAL 100MG TABLET", "است", "30300", "70%", "48700", "79000"])
        insurance.save(path)

    def test_number_parser_preserves_decimals_and_thousands(self) -> None:
        self.assertAlmostEqual(number("71.693"), 71.693)
        self.assertAlmostEqual(number("0.1290055350553505"), 0.1290055350553505)
        self.assertEqual(number("22,225"), 22_225)
        self.assertEqual(number("۷۰٪"), 70)

    def test_rial_price_is_converted_to_toman(self) -> None:
        headers = ["قیمت مصرف کننده ریال"]
        price, warning = price_payload(
            {headers[0]: 1_230_000},
            headers,
            "https://example.test",
            "2026-08-04T00:00:00+00:00",
            None,
        )
        self.assertIsNone(warning)
        self.assertEqual(price["amountToman"], 123_000)
        self.assertEqual(price["sourceCurrency"], "IRR")

    def test_real_provider_coverage_mappings(self) -> None:
        health_headers = ["کد ژنريک", "کد برند", "عنوان", " درصد سهم سازمان با احتساب يارانه ارزي"]
        health_row = {
            "کد ژنريک": "02015",
            "کد برند": "02015",
            "عنوان": "ACARBOSE",
            " درصد سهم سازمان با احتساب يارانه ارزي": "88.493",
        }
        coverage, warning = coverage_payload(health_row, health_headers, "health_insurance", "u", "t")
        self.assertIsNone(warning)
        self.assertAlmostEqual(coverage["percent"], 88.493)
        self.assertEqual(coverage["genericCode"], "02015")

        armed_headers = ["کد ژنريک", "نام ژنريك دارو", "درصد سهم بيمار عادي با احتساب يارانه"]
        armed_row = {
            "کد ژنريک": "02015",
            "نام ژنريك دارو": "ACARBOSE 100 MG TABLET ORAL TABLET 02015",
            "درصد سهم بيمار عادي با احتساب يارانه": 0.1290055350553505,
        }
        coverage, warning = coverage_payload(armed_row, armed_headers, "armed_forces", "u", "t")
        self.assertIsNone(warning)
        self.assertAlmostEqual(coverage["percent"], 87.099446, places=6)

        social_headers = ["کد دارو", "نام دارو", "درصد سازمان از قیمت بدون یارانه"]
        social_row = {
            "کد دارو": "02015",
            "نام دارو": "ACARBOSE 100 MG TABLET ORAL",
            "درصد سازمان از قیمت بدون یارانه": "70%",
        }
        coverage, warning = coverage_payload(social_row, social_headers, "social_security", "u", "t")
        self.assertIsNone(warning)
        self.assertEqual(coverage["percent"], 70)

    def test_complete_four_source_bundle_with_real_headers_is_publishable(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            scope_path = root / "scope.json"
            self.write_scope(scope_path)
            nfi_path = root / "nfi.xlsx"
            insurance_path = root / "insurance.xlsx"

            nfi = Workbook()
            products = nfi.active
            products.title = "NFI_Products_1"
            products.append([
                "Generic_INN", "NFI_Detail_ID", "NFI_Title_FA", "NFI_Generic_Code",
                "NFI_IRC_Code", "Dosage_Form", "Strength", "Price_Per_Pack_IRR",
                "Match_Confidence_0_100",
            ])
            products.append([
                "Acarbose", "1", "ACARBOSE TEST TABLET 100 mg", "NFI-02015",
                "6260001", "TABLET", "100 mg", 1_200_000, 99,
            ])
            nfi.save(nfi_path)
            self.write_realistic_insurance_workbook(insurance_path)

            bundle = build_bundle(nfi_path, insurance_path, None, scope_path)
            statuses = {source["sourceId"]: source["status"] for source in bundle["run"]["sources"]}
            self.assertEqual(statuses, {
                "iran_fda_nfi": "succeeded",
                "health_insurance": "succeeded",
                "armed_forces": "succeeded",
                "social_security": "succeeded",
            })
            self.assertEqual(bundle["run"]["status"], "ready_to_publish")
            insurer_records = [record for record in bundle["records"] if record["insuranceCoverages"]]
            self.assertEqual(len(insurer_records), 3)
            self.assertTrue(all("price" not in record for record in insurer_records))
            percentages = {
                record["insuranceCoverages"][0]["provider"]: record["insuranceCoverages"][0]["percent"]
                for record in insurer_records
            }
            self.assertAlmostEqual(percentages["health_insurance"], 88.493)
            self.assertAlmostEqual(percentages["armed_forces"], 87.099446, places=6)
            self.assertEqual(percentages["social_security"], 70)

    def test_rich_nfi_products_sheet_is_detected_and_preserves_review_confidence(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            scope_path = root / "scope.json"
            self.write_scope(scope_path)
            nfi_path = root / "nfi.xlsx"
            insurance_path = root / "insurance.xlsx"

            nfi = Workbook()
            summary = nfi.active
            summary.title = "NFI_Match_Summary"
            summary.append(["Drug_ID", "Generic_INN"])
            summary.append(["WD-1", "Metformin"])
            products = nfi.create_sheet("NFI_Products_1")
            products.append([
                "Generic_INN", "Match_Confidence_0_100", "NFI_Detail_ID", "NFI_Product_URL",
                "NFI_Title_FA", "NFI_Generic_Code", "NFI_IRC_Code", "Dosage_Form", "Strength",
                "Price_Per_Pack_IRR", "All_Label_Value_JSON", "Raw_Detail_Text",
            ])
            products.append([
                "Metformin", 96, "101", "https://irc.fda.gov.ir/NFI/Detail/101",
                "متفورمین تست قرص خوراکی 500 mg", "NFI-G-1", None, "TABLET", "500 mg",
                1_200_000, '{"نام":"METFORMIN TEST"}', ":IRC 6260000000001 :GTIN 06260000000001",
            ])
            products.append([
                "Metformin", 91, "101", "https://irc.fda.gov.ir/NFI/Detail/101",
                "متفورمین تست قرص خوراکی 500 mg", "NFI-G-1", None, "TABLET", "500 mg",
                1_200_000, '{"نام":"METFORMIN TEST"}', ":IRC 6260000000001 :GTIN 06260000000001",
            ])
            products.append([
                "Metformin", 85, "102", "https://irc.fda.gov.ir/NFI/Detail/102",
                "متفورمین نیازمند بازبینی قرص خوراکی 1000 mg", "NFI-G-1", "6260000000002",
                "TABLET", "1000 mg", 2_000_000, None, "",
            ])
            nfi.save(nfi_path)
            self.write_realistic_insurance_workbook(insurance_path, "Metformin", "00001")

            bundle = build_bundle(nfi_path, insurance_path, None, scope_path)
            nfi_records = [record for record in bundle["records"] if not record["insuranceCoverages"]]
            self.assertEqual(len(nfi_records), 2)
            verified = next(record for record in nfi_records if record["matchConfidence"] == 0.96)
            review = next(record for record in nfi_records if record["matchConfidence"] == 0.85)
            self.assertEqual(verified["brandName"], "METFORMIN TEST")
            self.assertEqual(verified["brandRegistryCode"], "6260000000001")
            self.assertEqual(verified["price"]["amountToman"], 120_000)
            self.assertIn("GTIN 06260000000001", verified["sourceReference"])
            self.assertEqual(review["brandName"], "متفورمین نیازمند بازبینی")
            self.assertEqual(bundle["run"]["status"], "needs_review")
            self.assertEqual(bundle["run"]["summary"]["ambiguousMatchCount"], 1)
            self.assertTrue(any("1 ردیف تکراری حذف" in item for item in bundle["diagnostics"]))


if __name__ == "__main__":
    unittest.main()
