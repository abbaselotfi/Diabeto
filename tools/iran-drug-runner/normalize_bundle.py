from __future__ import annotations

import hashlib
import json
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_META = {
    "health_insurance": {
        "sheet": "بیمه سلامت",
        "url": "https://mdp.ihio.gov.ir/",
        "generic_name": ["عنوان"],
        "generic_code": ["کد ژنریک", "کد ژنريک"],
        "brand_code": ["کد برند"],
        "dosage_form": ["شکل"],
        "strength": ["دوز"],
        "coverage_percent": [
            "درصد سهم سازمان با احتساب یارانه ارزی",
            "درصد سهم سازمان با احتساب يارانه ارزي",
        ],
        "coverage_mode": "organization_percent",
    },
    "armed_forces": {
        "sheet": "ساتا",
        "url": "https://esata.ir/web/sakhad/drug/",
        "generic_name": ["نام ژنريك دارو", "نام ژنریک دارو"],
        "generic_code": ["کد ژنریک", "کد ژنريک"],
        "brand_code": [],
        "dosage_form": [],
        "strength": [],
        "patient_percent": [
            "درصد سهم بيمار عادي با احتساب يارانه",
            "درصد سهم بیمار عادی با احتساب یارانه",
        ],
        "coverage_mode": "patient_percent_complement",
    },
    "social_security": {
        "sheet": "تامین اجتماعی",
        "url": "https://darman.tamin.ir/Forms/Public/Druglist.aspx?pagename=hdpDrugList",
        "generic_name": ["نام دارو"],
        "generic_code": ["کد دارو"],
        "brand_code": [],
        "dosage_form": [],
        "strength": [],
        "coverage_percent": ["درصد سازمان از قیمت بدون یارانه"],
        "coverage_mode": "organization_percent",
    },
}

DEFAULT_SCOPE_PATH = Path(__file__).resolve().with_name("scope_allowlist.json")

# NFI / legacy aliases only. Insurance sheets use SOURCE_META exact mappings to
# avoid dangerous fuzzy matches between similarly named financial columns.
HEADER_ALIASES = {
    "generic_name": ["نام ژنریک", "نام علمی", "generic inn", "generic name", "genericname", "نام دارو"],
    "brand_name": ["نام برند", "نام تجاری", "nfi title fa", "brand name", "brandname"],
    "generic_code": ["کد ژنریک", "nfi generic code", "generic code", "genericcode", "کد دارو"],
    "brand_code": ["کد برند", "nfi irc code", "nfi product code", "کد تجاری", "brand code", "brandcode", "کد فرآورده", "irc"],
    "dosage_form": ["شکل دارویی", "dosage form", "dosageform", "فرم دارویی"],
    "strength": ["قدرت", "دوز", "strength", "strength presentation"],
    "price": ["قیمت مصرف کننده", "قیمت مصرف‌کننده", "price per pack irr", "consumer price", "price"],
    "source_link": ["لینک", "آدرس منبع", "nfi product url", "source url", "url"],
}

NFI_PRODUCTS_SHEET_PREFIX = "nfi products"
NFI_CONFIDENCE_HEADER = "Match_Confidence_0_100"
NFI_DETAIL_ID_HEADER = "NFI_Detail_ID"
NFI_TITLE_HEADER = "NFI_Title_FA"
NFI_BRAND_OWNER_HEADER = "NFI_Brand_Name"
NFI_LABELS_JSON_HEADER = "All_Label_Value_JSON"
NFI_RAW_DETAIL_HEADER = "Raw_Detail_Text"
NFI_GTIN_HEADER = "GTIN"

DIGIT_TRANSLATION = str.maketrans(
    "يىك۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩٫٬",
    "ییک01234567890123456789.,",
)


def normalize_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.translate(DIGIT_TRANSLATION)
    text = text.replace("\u200c", " ")
    return re.sub(r"[^a-z0-9آ-ی]+", " ", text).strip()


def number(value: Any) -> float | None:
    """Parse Iranian/English numeric cells without destroying decimal points."""
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().translate(DIGIT_TRANSLATION)
    if not text:
        return None
    text = text.replace("٪", "").replace("%", "")
    text = text.replace("\u200c", "").replace(" ", "")
    # Source spreadsheets use comma/Arabic-thousands separators as grouping.
    text = text.replace(",", "")
    text = re.sub(r"[^0-9+\-.]", "", text)
    if text in {"", "+", "-", ".", "+.", "-."}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_percent(value: Any, *, fraction_when_unit_interval: bool = False) -> float | None:
    raw = number(value)
    if raw is None:
        return None
    text = str(value or "")
    has_percent_sign = "%" in text or "٪" in text
    if fraction_when_unit_interval and not has_percent_sign and 0 <= raw <= 1:
        raw *= 100
    return raw


def load_scope(path: Path) -> list[dict[str, Any]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    entries = payload.get("entries")
    if payload.get("schemaVersion") != 1 or not isinstance(entries, list) or not entries:
        raise ValueError("فهرست مجاز دامنه دارویی معتبر یا غیرخالی نیست.")
    validated: list[dict[str, Any]] = []
    for entry in entries:
        canonical = str(entry.get("canonicalName") or "").strip()
        aliases = [str(alias).strip() for alias in entry.get("aliases", []) if str(alias).strip()]
        domains = [str(domain).strip() for domain in entry.get("clinicalDomains", []) if str(domain).strip()]
        if not canonical or not aliases or not domains:
            raise ValueError(f"ردیف ناقص در فهرست مجاز: {entry}")
        validated.append({"canonicalName": canonical, "aliases": aliases, "clinicalDomains": domains})
    return validated


def match_scope(name: str, scope: list[dict[str, Any]]) -> dict[str, Any] | None:
    normalized_name = normalize_text(name)
    if not normalized_name:
        return None
    exact: list[dict[str, Any]] = []
    contained: list[dict[str, Any]] = []
    for entry in scope:
        aliases = {normalize_text(alias) for alias in entry["aliases"]}
        if normalized_name in aliases:
            exact.append(entry)
            continue
        if any(len(alias) >= 5 and alias in normalized_name for alias in aliases):
            contained.append(entry)
    token_matches: list[dict[str, Any]] = []
    if not exact and not contained:
        name_tokens = set(normalized_name.split())
        if len(name_tokens) >= 2:
            for entry in scope:
                if any(set(normalize_text(alias).split()) == name_tokens for alias in entry["aliases"]):
                    token_matches.append(entry)
    candidates = exact or contained or token_matches
    unique = {entry["canonicalName"]: entry for entry in candidates}
    return next(iter(unique.values())) if len(unique) == 1 else None


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def workbook_rows(path: Path, sheet_name: str | None = None) -> tuple[list[dict[str, Any]], list[str], str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, ())]
        rows = [dict(zip(headers, values)) for values in iterator if any(value not in (None, "") for value in values)]
        return rows, headers, sheet.title
    finally:
        workbook.close()


def nfi_workbook_rows(path: Path) -> tuple[list[dict[str, Any]], list[str], str]:
    """Read all NFI_Products* sheets, falling back to the first sheet for legacy files."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        product_sheets = [
            sheet
            for sheet in workbook.worksheets
            if normalize_text(sheet.title).startswith(NFI_PRODUCTS_SHEET_PREFIX)
        ]
        selected = product_sheets or [workbook[workbook.sheetnames[0]]]
        rows: list[dict[str, Any]] = []
        combined_headers: list[str] = []
        for sheet in selected:
            iterator = sheet.iter_rows(values_only=True)
            headers = [str(item or "").strip() for item in next(iterator, ())]
            for header in headers:
                if header and header not in combined_headers:
                    combined_headers.append(header)
            rows.extend(
                dict(zip(headers, values))
                for values in iterator
                if any(item not in (None, "") for item in values)
            )
        return rows, combined_headers, ", ".join(sheet.title for sheet in selected)
    finally:
        workbook.close()


def direct_value(row: dict[str, Any], header: str) -> Any:
    wanted = normalize_text(header)
    for current, item in row.items():
        if normalize_text(current) == wanted:
            return item
    return None


def exact_header(headers: list[str], aliases: list[str]) -> str | None:
    normalized = {normalize_text(header): header for header in headers}
    for alias in aliases:
        found = normalized.get(normalize_text(alias))
        if found is not None:
            return found
    return None


def provider_header(headers: list[str], provider: str, field: str) -> str | None:
    aliases = SOURCE_META[provider].get(field, [])
    return exact_header(headers, aliases) if isinstance(aliases, list) else None


def provider_value(row: dict[str, Any], headers: list[str], provider: str, field: str) -> Any:
    header = provider_header(headers, provider, field)
    return row.get(header) if header else None


def validate_provider_headers(headers: list[str], provider: str) -> None:
    meta = SOURCE_META[provider]
    missing: list[str] = []
    if not provider_header(headers, provider, "generic_name"):
        missing.append("نام دارو")
    mode = meta["coverage_mode"]
    if mode == "organization_percent" and not provider_header(headers, provider, "coverage_percent"):
        missing.append("درصد سهم سازمان")
    if mode == "patient_percent_complement" and not provider_header(headers, provider, "patient_percent"):
        missing.append("درصد سهم بیمار")
    if missing:
        available = " | ".join(header for header in headers if header)[:1200]
        raise ValueError(f"ستون‌های ضروری {', '.join(missing)} شناخته نشد. هدرهای موجود: {available}")


def nfi_match_confidence(row: dict[str, Any]) -> float:
    raw = number(direct_value(row, NFI_CONFIDENCE_HEADER))
    if raw is None:
        return 1
    fraction = raw / 100 if raw > 1 else raw
    return max(0, min(1, fraction))


def nfi_json_labels(row: dict[str, Any]) -> dict[str, Any]:
    raw = direct_value(row, NFI_LABELS_JSON_HEADER)
    if not isinstance(raw, str) or not raw.strip():
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def header_for(headers: list[str], field: str) -> str | None:
    normalized = {normalize_text(header): header for header in headers}
    for alias in HEADER_ALIASES[field]:
        exact = normalized.get(normalize_text(alias))
        if exact is not None:
            return exact
    for alias in HEADER_ALIASES[field]:
        normalized_alias = normalize_text(alias)
        for normalized_header, original_header in normalized.items():
            if normalized_alias and normalized_alias in normalized_header:
                return original_header
    return None


def value(row: dict[str, Any], headers: list[str], field: str) -> Any:
    header = header_for(headers, field)
    return row.get(header) if header else None


def nfi_registry_code(row: dict[str, Any], headers: list[str]) -> str | None:
    direct = str(value(row, headers, "brand_code") or "").strip()
    if direct:
        return direct
    labels = nfi_json_labels(row)
    for key in ("IRC", "کد IRC", "کد فرآورده"):
        candidate = str(labels.get(key) or "").strip()
        if candidate:
            return candidate
    raw = str(direct_value(row, NFI_RAW_DETAIL_HEADER) or "")
    match = re.search(r"(?:\bIRC\b\s*:|:IRC\b)\s*([0-9]{6,})", raw, flags=re.IGNORECASE)
    return match.group(1) if match else None


def nfi_gtin(row: dict[str, Any]) -> str | None:
    direct = str(direct_value(row, NFI_GTIN_HEADER) or "").strip()
    if direct:
        return direct
    labels = nfi_json_labels(row)
    candidate = str(labels.get("GTIN") or "").strip()
    if candidate:
        return candidate
    raw = str(direct_value(row, NFI_RAW_DETAIL_HEADER) or "")
    match = re.search(r"(?:\bGTIN\b\s*:|:GTIN\b)\s*([0-9]{6,})", raw, flags=re.IGNORECASE)
    return match.group(1) if match else None


def nfi_brand_name(row: dict[str, Any], headers: list[str]) -> str | None:
    labels = nfi_json_labels(row)
    labelled_name = str(labels.get("نام") or labels.get("Name") or "").strip()
    if labelled_name:
        return labelled_name
    title = str(direct_value(row, NFI_TITLE_HEADER) or "").strip()
    if title:
        compact_title = title.split("(", 1)[0].strip()
        markers = (
            " قرص", " کپسول", " محلول", " سوسپانسیون", " تزریقی", " پودر",
            " شربت", " ویال", " آمپول", " قلم", " اسپری", " قطره", " کرم",
            " پماد", " ژل", " شیاف", " استنشاقی", " خوراکی", " پرنترال",
            " tablet", " capsule", " solution", " suspension", " injection",
            " powder", " syrup", " vial", " ampoule", " pen", " spray",
            " drops", " cream", " ointment", " gel", " suppository",
        )
        normalized_title = compact_title.lower()
        positions = [normalized_title.find(marker) for marker in markers]
        cut_at = min((position for position in positions if position > 0), default=-1)
        trade_name = compact_title[:cut_at].strip() if cut_at > 0 else compact_title
        return trade_name or title
    header = header_for(headers, "brand_name")
    if header and normalize_text(header) != normalize_text(NFI_BRAND_OWNER_HEADER):
        return str(row.get(header) or "").strip() or None
    return None


def nfi_record_identity(record: dict[str, Any], row: dict[str, Any]) -> tuple[str, ...]:
    registry_code = normalize_text(record.get("brandRegistryCode"))
    if registry_code:
        return ("irc", registry_code)
    gtin = normalize_text(nfi_gtin(row))
    if gtin:
        return ("gtin", gtin)
    detail_id = normalize_text(direct_value(row, NFI_DETAIL_ID_HEADER))
    if detail_id:
        return ("detail", detail_id)
    return (
        "label",
        normalize_text(record.get("genericName")),
        normalize_text(record.get("brandName")),
        normalize_text(record.get("dosageForm")),
        normalize_text(record.get("strengthPresentation")),
    )


def price_currency(header: str | None, default_currency: str | None) -> str | None:
    normalized = normalize_text(header)
    if "ریال" in normalized or "rial" in normalized or "irr" in normalized.split():
        return "IRR"
    if "تومان" in normalized or "toman" in normalized.split():
        return "TOMAN"
    if default_currency in {"IRR", "TOMAN"}:
        return default_currency
    return None


def price_payload(
    row: dict[str, Any],
    headers: list[str],
    source_url: str,
    observed_at: str,
    default_currency: str | None,
) -> tuple[dict[str, Any] | None, str | None]:
    """NFI consumer price only. Insurance prices are intentionally not mapped here."""
    header = header_for(headers, "price")
    amount = number(row.get(header)) if header else None
    if amount is None:
        return None, None
    currency = price_currency(header, default_currency)
    if not currency:
        return None, f"واحد پول ستون «{header}» مشخص نیست؛ IRR یا TOMAN باید صریح باشد."
    amount_toman = round(amount / 10 if currency == "IRR" else amount)
    return {
        "amountToman": amount_toman,
        "priceKind": "consumer_retail",
        "sourceAmount": amount,
        "sourceCurrency": currency,
        "effectiveAt": observed_at,
        "sourceUrl": source_url,
        "sourceReference": header,
    }, None


def coverage_payload(
    row: dict[str, Any],
    headers: list[str],
    provider: str,
    source_url: str,
    observed_at: str,
    default_currency: str | None = None,
) -> tuple[dict[str, Any] | None, str | None]:
    del default_currency
    meta = SOURCE_META[provider]
    mode = meta["coverage_mode"]
    source_reference = ""

    if mode == "organization_percent":
        header = provider_header(headers, provider, "coverage_percent")
        percent = normalize_percent(row.get(header) if header else None)
        source_reference = header or ""
    elif mode == "patient_percent_complement":
        header = provider_header(headers, provider, "patient_percent")
        patient_percent = normalize_percent(
            row.get(header) if header else None,
            fraction_when_unit_interval=True,
        )
        percent = None if patient_percent is None else 100 - patient_percent
        source_reference = header or ""
    else:
        return None, f"روش محاسبه پوشش برای منبع {provider} تعریف نشده است."

    if percent is None:
        return None, "درصد پوشش بیمه از ستون رسمی منبع قابل محاسبه نبود."
    percent = round(percent, 6)
    if not 0 <= percent <= 100:
        return None, f"درصد پوشش خارج از محدوده است: {percent}"

    return {
        "provider": provider,
        "percent": percent,
        "origin": "source",
        "genericCode": str(provider_value(row, headers, provider, "generic_code") or "").strip() or None,
        "brandCode": str(provider_value(row, headers, provider, "brand_code") or "").strip() or None,
        "effectiveAt": observed_at,
        "sourceUrl": source_url,
        "sourceReference": source_reference or "خروجی رسمی بیمه",
    }, None


def make_record(
    row: dict[str, Any],
    headers: list[str],
    source_url: str,
    observed_at: str,
    scope_entry: dict[str, Any],
) -> dict[str, Any] | None:
    generic_name = str(value(row, headers, "generic_name") or "").strip()
    brand_name = nfi_brand_name(row, headers)
    if not generic_name:
        return None
    row_url = str(value(row, headers, "source_link") or source_url).strip()
    return {
        "genericName": scope_entry["canonicalName"],
        "genericRegistryCode": str(value(row, headers, "generic_code") or "").strip() or None,
        "brandName": brand_name,
        "brandRegistryCode": nfi_registry_code(row, headers),
        "dosageForm": str(value(row, headers, "dosage_form") or "").strip() or None,
        "strengthPresentation": str(value(row, headers, "strength") or "").strip() or None,
        "clinicalDomains": scope_entry["clinicalDomains"],
        "insuranceCoverages": [],
        "sourceUrl": row_url,
        "sourceReference": " · ".join(filter(None, [
            f"Iran FDA NFI · {generic_name}",
            f"Detail {str(direct_value(row, NFI_DETAIL_ID_HEADER) or '').strip()}" if direct_value(row, NFI_DETAIL_ID_HEADER) else None,
            f"GTIN {nfi_gtin(row)}" if nfi_gtin(row) else None,
        ])),
        "observedAt": observed_at,
        "matchConfidence": nfi_match_confidence(row),
    }


def insurance_record_fields(row: dict[str, Any], headers: list[str], provider: str) -> tuple[str | None, str | None]:
    dosage_form = str(provider_value(row, headers, provider, "dosage_form") or "").strip() or None
    strength = str(provider_value(row, headers, provider, "strength") or "").strip() or None
    return dosage_form, strength


def build_bundle(
    nfi_path: Path,
    insurance_path: Path,
    default_currency: str | None = None,
    scope_path: Path = DEFAULT_SCOPE_PATH,
) -> dict[str, Any]:
    run_id = str(uuid.uuid4())
    observed_at = datetime.now(timezone.utc).isoformat()
    source_runs: list[dict[str, Any]] = []
    records: list[dict[str, Any]] = []
    errors: list[str] = []
    diagnostics: list[str] = []

    try:
        scope = load_scope(scope_path)
    except Exception as exc:
        scope = []
        errors.append(f"Scope: {exc}")

    nfi_errors: list[str] = []
    try:
        nfi_rows, nfi_headers, nfi_sheet = nfi_workbook_rows(nfi_path)
        if not header_for(nfi_headers, "generic_name"):
            raise ValueError("ستون نام ژنریک در خروجی NFI شناخته نشد.")
        skipped_outside_scope = 0
        duplicate_rows = 0
        seen_products: dict[tuple[str, ...], int] = {}
        for row in nfi_rows:
            raw_generic_name = str(value(row, nfi_headers, "generic_name") or "").strip()
            scope_entry = match_scope(raw_generic_name, scope)
            if not scope_entry:
                skipped_outside_scope += 1
                continue
            record = make_record(
                row,
                nfi_headers,
                "https://irc.fda.gov.ir/nfi",
                observed_at,
                scope_entry,
            )
            if not record:
                continue
            price, warning = price_payload(row, nfi_headers, record["sourceUrl"], observed_at, default_currency)
            if warning:
                nfi_errors.append(warning)
            if price:
                record["price"] = price
            identity = nfi_record_identity(record, row)
            existing_index = seen_products.get(identity)
            if existing_index is not None:
                duplicate_rows += 1
                if record["matchConfidence"] > records[existing_index].get("matchConfidence", 0):
                    records[existing_index] = record
                continue
            seen_products[identity] = len(records)
            records.append(record)
        if not records:
            raise ValueError("هیچ ردیف NFI با فهرست مجاز دامنه دارویی تطبیق پیدا نکرد.")
        diagnostics.append(
            f"NFI ({nfi_sheet}): {len(seen_products)} فرآورده وارد، {duplicate_rows} ردیف تکراری حذف و {skipped_outside_scope} ردیف خارج از دامنه رد شد."
        )
        source_runs.append({
            "sourceId": "iran_fda_nfi",
            "status": "succeeded" if not nfi_errors else "needs_review",
            "rowCount": len(nfi_rows),
            "startedAt": observed_at,
            "completedAt": observed_at,
            "sourceUrl": "https://irc.fda.gov.ir/nfi",
            "checksumSha256": sha256(nfi_path),
            "error": nfi_errors[0] if nfi_errors else None,
        })
        errors.extend(f"NFI: {message}" for message in nfi_errors[:100])
    except Exception as exc:
        errors.append(f"NFI: {exc}")
        source_runs.append({
            "sourceId": "iran_fda_nfi",
            "status": "failed",
            "startedAt": observed_at,
            "completedAt": observed_at,
            "sourceUrl": "https://irc.fda.gov.ir/nfi",
            "error": str(exc),
        })

    nfi_names = {normalize_text(record["genericName"]): record for record in records}
    for provider, meta in SOURCE_META.items():
        source_errors: list[str] = []
        try:
            rows, headers, sheet = workbook_rows(insurance_path, meta["sheet"])
            validate_provider_headers(headers, provider)
            emitted = 0
            skipped_outside_scope = 0
            for row_number, row in enumerate(rows, start=2):
                raw_name = str(provider_value(row, headers, provider, "generic_name") or "").strip()
                scope_entry = match_scope(raw_name, scope)
                if not scope_entry:
                    skipped_outside_scope += 1
                    continue

                normalized = normalize_text(raw_name)
                exact = nfi_names.get(normalized)
                candidates = [
                    entry for key, entry in nfi_names.items()
                    if key and (key in normalized or normalized in key)
                ] if not exact else [exact]
                base = exact or (candidates[0] if len(candidates) == 1 else None)

                coverage, coverage_error = coverage_payload(
                    row,
                    headers,
                    provider,
                    meta["url"],
                    observed_at,
                    default_currency,
                )
                if coverage_error:
                    source_errors.append(f"ردیف {row_number}: {coverage_error}")
                    continue

                dosage_form, strength = insurance_record_fields(row, headers, provider)
                record = {
                    "genericName": base["genericName"] if base else scope_entry["canonicalName"],
                    "genericRegistryCode": base.get("genericRegistryCode") if base else None,
                    "brandName": None,
                    "brandRegistryCode": None,
                    "dosageForm": dosage_form,
                    "strengthPresentation": strength,
                    "clinicalDomains": scope_entry["clinicalDomains"],
                    "insuranceCoverages": [coverage] if coverage else [],
                    "sourceUrl": meta["url"],
                    "sourceReference": f"خروجی {meta['sheet']}",
                    "observedAt": observed_at,
                    "matchConfidence": 0.99 if exact else 0.92 if base else 0.4,
                }
                records.append({key: item for key, item in record.items() if item is not None})
                emitted += 1

            diagnostics.append(
                f"{meta['sheet']}: {emitted} ردیف وارد و {skipped_outside_scope} ردیف خارج از دامنه رد شد."
            )
            status = "succeeded" if not source_errors else "needs_review"
            source_runs.append({
                "sourceId": provider,
                "status": status,
                "rowCount": len(rows),
                "startedAt": observed_at,
                "completedAt": observed_at,
                "sourceUrl": meta["url"],
                "checksumSha256": sha256(insurance_path),
                "error": source_errors[0] if source_errors else None,
            })
            errors.extend(f"{meta['sheet']}: {message}" for message in source_errors[:100])
        except Exception as exc:
            errors.append(f"{meta['sheet']}: {exc}")
            source_runs.append({
                "sourceId": provider,
                "status": "failed",
                "startedAt": observed_at,
                "completedAt": observed_at,
                "sourceUrl": meta["url"],
                "error": str(exc),
            })

    brand_count = len({
        (normalize_text(record["genericName"]), normalize_text(record.get("brandName")))
        for record in records if record.get("brandName")
    })
    generic_count = len({normalize_text(record["genericName"]) for record in records})
    ambiguous_count = sum(1 for record in records if record.get("matchConfidence", 0) < 0.9)
    run_status = "ready_to_publish" if all(source["status"] == "succeeded" for source in source_runs) and not ambiguous_count else "needs_review"
    return {
        "schemaVersion": 1,
        "run": {
            "id": run_id,
            "schemaVersion": 1,
            "status": run_status,
            "startedAt": observed_at,
            "completedAt": datetime.now(timezone.utc).isoformat(),
            "sources": source_runs,
            "summary": {
                "genericCount": generic_count,
                "brandCount": brand_count,
                "priceChangeCount": sum(1 for record in records if record.get("price")),
                "coverageChangeCount": sum(len(record.get("insuranceCoverages", [])) for record in records),
                "ambiguousMatchCount": ambiguous_count,
                "errorCount": len(errors),
            },
        },
        "records": records,
        "diagnostics": diagnostics + errors,
    }


def write_bundle(
    nfi_path: Path,
    insurance_path: Path,
    output_path: Path,
    default_currency: str | None = None,
    scope_path: Path = DEFAULT_SCOPE_PATH,
) -> dict[str, Any]:
    bundle = build_bundle(nfi_path, insurance_path, default_currency, scope_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(bundle, ensure_ascii=False, indent=2), encoding="utf-8")
    return bundle
